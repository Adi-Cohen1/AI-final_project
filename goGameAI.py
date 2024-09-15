import os
import sys
import random
from typing import Optional, Tuple
import tkinter as tk
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from GoBoard import GoBoard
from goDisplay import GoDisplay
from MCTS import MCTS
from Expectimax import Expectimax
from Minimax import Minimax
from MinimaxAlphaBeta import MinimaxAlphaBeta
from Agents import RandomAgent, GreedyAgent
from QLearning import QLearning

class GoGame:
    """
     Manages a Go game, including game setup, playing, strategy application, and result recording.

     Attributes:
         size (int): The size of the Go board (size x size).
         num_games (int): The number of games to be played.
         black_strategy (str): The strategy used by the black player.
         white_strategy (str): The strategy used by the white player.
         display (GoDisplay): An instance of GoDisplay for rendering the game.
         is_display (bool): Flag to determine if the game display should be shown.
         results (list): List of game results.
         current_game (int): The index of the current game being played.
         current_color (str): The color of the current player ('BLACK' or 'WHITE').
         game_over (bool): Flag indicating if the current game is over.
         finished (bool): Flag indicating if all games are finished.
         board (Optional[GoBoard]): The current game board instance.
         previous_boards (set): Set of previously seen board configurations.
         speed (int): The speed of the game (delay between moves in milliseconds).
         first_turn (bool): Flag indicating if it is the first turn of a game.
         black_strategy (str): Strategy name for the black player.
         white_strategy (str): Strategy name for the white player.
         random_agent_white (RandomAgent): Random agent for the white player.
         random_agent_black (RandomAgent): Random agent for the black player.
         greedy_agent_white (GreedyAgent): Greedy agent for the white player.
         greedy_agent_black (GreedyAgent): Greedy agent for the black player.
         qlearn_agent_black (QLearning): Q-learning agent for the black player.
         results_file (str): The file name for saving game results.
         wb (Workbook): The Excel workbook instance for saving results.
         ws (Worksheet): The worksheet instance for saving game results.
         black_strategies (dict): Dictionary mapping strategy names to strategy functions for black player.
         white_strategies (dict): Dictionary mapping strategy names to strategy functions for white player.
     """
    def __init__(self, size, num_games, black_strategy, while_strategy, display, is_display, save_in_excel=False):
        """
          Initialize the GoGame class with board size, strategies, display, and agents.

          Args:
              size (int): Size of the Go board.
              num_games (int): Number of games to play.
              black_strategy (str): The strategy name for the black player.
              white_strategy (str): The strategy name for the white player.
              display (GoDisplay): Display instance to visualize the game.
              is_display (bool): Whether to show the game board.
          """
        self.size = size
        self.num_games = num_games
        self.display = display
        if not is_display:
            self.display.root.withdraw()  # Hide the root window
        self.save_in_excel = save_in_excel
        self.results = []
        self.current_game = 0
        self.current_color = 'BLACK'
        self.game_over = False
        self.finished = False
        self.board = None
        self.previous_boards = set()
        self.speed = 1
        self.first_turn = True
        self.black_strategy = black_strategy
        self.white_strategy = while_strategy
        self.random_agent_white = RandomAgent('WHITE')
        self.random_agent_black = RandomAgent('BLACK')
        self.greedy_agent_white = GreedyAgent('WHITE')
        self.greedy_agent_black = GreedyAgent('BLACK')
        self.qlearn_agent_black = QLearning(exploration_rate=0.0)
        self.qlearn_agent_black.load("q_table_against_q_table_diff_heuristic.npy")

        # Initialize Excel Workbook and sheet
        if self.save_in_excel:
            self.results_file = f"{self.black_strategy} vs. {self.white_strategy}.xlsx"
            if os.path.exists(self.results_file):
                self.wb = load_workbook(self.results_file)
                self.ws = self.wb.active
            else:
                self.wb = Workbook()  # Create a new workbook if it doesn't exist
                self.ws = self.wb.active
                self.ws.title = "Game Results"
                self.ws.append(["game_num", "black_score", "white_score", "black_wins", "white_wins","tie"])  # Adding headers

        # Define a dictionary of strategies for BLACK
        self.black_strategies = {
            "random": self.random_agent_black.getAction,
            "greedy": self.greedy_agent_black.getAction,
            "monte_carlo": lambda board: self.monte_carlo_strategy(),
            "expectimax": lambda board: self.expectimax_strategy(),
            "minimax": lambda board: self.minimax_strategy(),
            "alpha_beta": lambda board: self.alpha_beta_strategy(),
            "qlearn": lambda board: self.qlearn_strategy(),
        }

        # Define a dictionary of strategies for WHITE
        self.white_strategies = {
            "random": self.random_agent_white.getAction,
            "greedy": self.greedy_agent_white.getAction,
        }

    def is_game_over(self) -> bool:
        """
          Check if the game is over by determining if either player can make a legal move.

          Returns:
              bool: True if no legal moves for both players, otherwise False.
          """
        black_moves = any(self.board.is_legal_move(x, y, 'BLACK') for x in range(self.size) for y in range(self.size))
        white_moves = any(self.board.is_legal_move(x, y, 'WHITE') for x in range(self.size) for y in range(self.size))
        return not black_moves and not white_moves

    def play_game_step(self):
        """
          Execute a single step in the game, including determining the move for the current player,
          updating the board, and checking for game termination.
          """
        if self.finished:
            print("game over")
            return

        if self.game_over:
            self.end_game()
            return

        # Initialize the board if it's the first step
        if self.board is None:
            self.board = GoBoard(self.size, self.previous_boards)
            self.current_color = 'BLACK'
            self.display.display_board(self.board)
            self.display.root.after(self.speed, self.play_game_step)
            return

        # Select the strategy based on the current color
        if self.current_color == 'BLACK':
            move = self.black_strategies[self.black_strategy](self.board)
        else:
            move = self.white_strategies[self.white_strategy](self.board)
        # Check if the game is over (both players passed)
        if move is None:
            if self.board:
                result = self.board.count_score()
                self.results.append(result)
                black_score = result["BLACK"]
                white_score = result["WHITE"]
                black_wins = 1 if black_score > white_score else 0
                white_wins = 1 if white_score > black_score else 0
                tie = 1 if white_score == black_score else 0

                if self.save_in_excel:
                    # Add the results to the Excel sheet
                    self.ws.append([self.current_game + 1, black_score, white_score, black_wins, white_wins, tie])

                print(f'Game {self.current_game + 1}: BLACK {black_score}, WHITE {white_score}')
                self.display.display_results(self.results)

            self.current_game += 1
            if self.current_game >= self.num_games:
                self.end_game()
                self.finished = True
            else:
                self.reset_game()
            return

        # Play the move and update the board
        x, y = move
        if self.board.play_actual_move(x, y, self.current_color):
            self.previous_boards.add(tuple(map(tuple, self.board.board)))
            self.current_color = 'WHITE' if self.current_color == 'BLACK' else 'BLACK'
            self.display.display_board(self.board)

        # Check if the game is over
        if self.is_game_over():
            self.game_over = True

        self.display.root.after(self.speed, self.play_game_step)

    def monte_carlo_strategy(self):
        """
         Apply the Monte Carlo Tree Search (MCTS) strategy to determine the next move.

         Returns:
             tuple: The chosen move (x, y) coordinates.
         """
        mcts = MCTS(self.board, self.current_color,self.greedy_agent_white, mcts_iterations=50, exploration_weight=1.5)
        move = mcts.mcts_search()
        return move

    def expectimax_strategy(self):
        """
        Apply the Expectimax algorithm to determine the next move.

        Returns:
            tuple: The chosen move (x, y) coordinates.
        """
        if self.first_turn:  # expectimax with depth=4 always return (3,0) for the first turn, so no need to check.
            move = (3, 0)
            self.first_turn = False
        else:
            expectimax_agent = Expectimax(self.board, self.current_color)
            move, value = expectimax_agent.expectimax(depth=4)
        return move

    def minimax_strategy(self):
        """
        Apply the Minimax algorithm to determine the next move.

        Returns:
            tuple: The chosen move (x, y) coordinates.
        """
        minimax_agent = Minimax(self.board, self.current_color)
        move, value = minimax_agent.minimax(depth=4)
        return move

    def alpha_beta_strategy(self):
        """
        Apply the Minimax algorithm with Alpha-Beta pruning to determine the next move.

        Returns:
            tuple: The chosen move (x, y) coordinates.
        """
        alpha_beta_agent = MinimaxAlphaBeta(self.board, self.current_color)
        move, value = alpha_beta_agent.minimax(depth=4)
        return move

    def qlearn_strategy(self):
        """
           Apply the Q-Learning agent's strategy to determine the next move.

           Returns:
               tuple: The chosen move (x, y) coordinates.
           """
        move = self.qlearn_agent_black.choose_action(self.board, self.current_color)
        return move

    def reset_game(self):
        """
        Reset the game state for a new game.
        """
        self.first_turn = True
        self.board = None
        self.current_color = 'BLACK'
        self.game_over = False
        self.previous_boards.clear()
        self.display.root.after(self.speed, self.play_game_step)

    def end_game(self):
        """
          End the game, save the results, and close the display window.
          """
        if self.board and self.current_game < self.num_games:
            result = self.board.count_score()
            self.results.append(result)
            black_score = result["BLACK"]
            white_score = result["WHITE"]
            black_wins = 1 if black_score > white_score else 0
            white_wins = 1 if white_score > black_score else 0
            tie = 1 if white_score == black_score else 0

            if self.save_in_excel:
                # Add the results to the Excel sheet
                self.ws.append([self.current_game + 1, black_score, white_score, black_wins, white_wins,tie])

            print(f'Game {self.current_game + 1} has ended: BLACK {black_score}, WHITE {white_score}')
            self.display.display_results(self.results)

        if self.current_game < self.num_games:
            self.current_game += 1
            self.reset_game()
        else:
            self.finished = True
            print(self.finished)
            # Save the Excel file after all games are completed
            if self.save_in_excel:
                self.wb.save(str(self.results_file))
                print(f"end game and results saved to {self.results_file}")

    def run(self):
        """
        Execute the game loop, playing one step of the game at a time.
        This function is the entry point to initiate the game sequence.
        """
        self.play_game_step()


if __name__ == "__main__":
    # Define the allowed strategies for black and white players
    black_strategies = ["random", "greedy", "minimax", "alpha_beta", "expectimax", "monte_carlo", "qlearn"]
    white_strategies = ["random", "greedy"]
    # Check if the command-line arguments are valid
    if len(sys.argv) != 5 \
            or sys.argv[1] not in black_strategies \
            or sys.argv[2] not in white_strategies \
            or not sys.argv[3].isdigit()\
            or int(sys.argv[3]) < 1\
            or sys.argv[4] not in ["display", "not_display"]:
        print("put in the command line: goGameAI.py "
              "<BLACK-STRATEGY> <WHITE-STRATEGY> <NUMBER-OF-GAME> <DISPLAY-BOARD (display | not_display)>")
        print("BLACK-STRATEGY: ", str(black_strategies))
        print("WHITE-STRATEGY: ", str(white_strategies))
        exit()
    # Game configuration based on command-line arguments
    size = 5
    black_strategy = sys.argv[1]
    while_strategy = sys.argv[2]
    num_games = int(sys.argv[3])
    is_display = sys.argv[4] == "display"

    root = tk.Tk()
    root.title("Go Game")
    # Create an instance of GoDisplay to handle the game's visual representation
    display = GoDisplay(root, size)
    # Initialize the Go game with the specified settings
    game = GoGame(size, num_games, black_strategy, while_strategy, display, is_display)
    print("start game")
    game.run()  # Start the game loop

    root.mainloop()  # Run the Tkinter event loop for the GUI
