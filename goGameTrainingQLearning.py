import os
import tkinter as tk
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from GoBoard import GoBoard
from goDisplay import GoDisplay
from QLearning import QLearning


class GoGameTrainingQLearning:
    def __init__(self, size: int, num_games: int, display, is_display):
        self.size = size
        self.num_games = num_games
        self.display = display
        if not is_display:
            self.display.root.withdraw()  # Hide the root window
        self.results = []
        self.current_game = 0
        self.current_color = 'BLACK'
        self.game_over = False
        self.finished = False
        self.board = None
        self.previous_boards = set()
        self.speed = 1
        self.first_turn = True
        self.qlearn_agent_black = QLearning()
        self.qlearn_agent_white = QLearning()

        # Initialize Excel Workbook and sheet
        self.results_file = "qlearn_vs_qlearn.xlsx"
        if os.path.exists(self.results_file):
            self.wb = load_workbook(self.results_file)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()  # Create a new workbook if it doesn't exist
            self.ws = self.wb.active
            self.ws.title = "Game Results"
            self.ws.append(["game_num", "black_score", "white_score", "black_wins", "white_wins", "tie"])  # Adding headers

    def is_game_over(self) -> bool:
        black_moves = any(self.board.is_legal_move(x, y, 'BLACK') for x in range(self.size) for y in range(self.size))
        white_moves = any(self.board.is_legal_move(x, y, 'WHITE') for x in range(self.size) for y in range(self.size))
        return not black_moves and not white_moves

    def play_game_step(self):
        if self.finished:
            print("game over")
            return

        if self.game_over:
            self.end_game()
            return

        if self.board is None:
            self.board = GoBoard(self.size, self.previous_boards)
            self.current_color = 'BLACK'
            self.display.display_board(self.board)
            self.display.root.after(self.speed, self.play_game_step)

            # initialize prev_board for BLACK and WHITE
            self.prev_black_board = self.board.copy()
            self.prev_white_board = self.board.copy()
            return

        # Select the strategy based on the current color
        if self.current_color == 'BLACK':
            move = self.qlearn_agent_black.choose_action(self.board, self.current_color)
        else:
            move = self.qlearn_agent_white.choose_action(self.board, self.current_color)

        if move is None:
            if self.board:
                result = self.board.count_score()
                self.results.append(result)
                black_score = result["BLACK"]
                white_score = result["WHITE"]
                black_wins = 1 if black_score > white_score else 0
                white_wins = 1 if white_score > black_score else 0
                tie = 1 if white_score == black_score else 0

                # Decrease the exploration rate at the end of every game
                self.qlearn_agent_black.decay_exploration_rate()
                self.qlearn_agent_white.decay_exploration_rate()

                # Add the results to the Excel sheet
                self.ws.append([self.current_game + 1, black_score, white_score, black_wins, white_wins, tie])

                print(f'Game {self.current_game + 1}: BLACK {black_score}, WHITE {white_score}')
                self.display.display_results(self.results)

            self.current_game += 1
            if self.current_game >= self.num_games:
                self.end_game()
                self.finished = True
                # Save the resulting Q-table
                self.qlearn_agent_black.save(f"q_table_against_q_table_{self.num_games}_iterations")
            else:
                self.reset_game()
            return

        x, y = move
        if self.board.play_actual_move(x, y, self.current_color):
            # Update the values in the Q-table
            if self.current_color == "BLACK":
                reward = self.qlearn_agent_black.get_reward(self.board, self.current_color)
                curr_board = self.board.copy()
                self.qlearn_agent_black.update_q_values(self.prev_black_board, self.current_color, move, reward, curr_board)
                self.prev_black_board = self.board.copy()
                if self.first_turn:
                    self.prev_white_board = self.board.copy()
                    self.first_turn = False
            else:  # current color is WHITE
                reward = self.qlearn_agent_white.get_reward(self.board, self.current_color)
                curr_board = self.board.copy()
                self.qlearn_agent_white.update_q_values(self.prev_white_board, self.current_color, move, reward, curr_board)
                self.prev_white_board = self.board.copy()

            self.previous_boards.add(tuple(map(tuple, self.board.board)))
            self.current_color = 'WHITE' if self.current_color == 'BLACK' else 'BLACK'
            self.display.display_board(self.board)

        if self.is_game_over():
            self.game_over = True

        self.display.root.after(self.speed, self.play_game_step)

    def reset_game(self):
        self.first_turn = True
        self.board = None
        self.current_color = 'BLACK'
        self.game_over = False
        self.previous_boards.clear()
        self.display.root.after(self.speed, self.play_game_step)

    def end_game(self):
        if self.board and self.current_game < self.num_games:
            result = self.board.count_score()
            self.results.append(result)
            black_score = result["BLACK"]
            white_score = result["WHITE"]
            black_wins = 1 if black_score > white_score else 0
            white_wins = 1 if white_score > black_score else 0
            tie = 1 if white_score == black_score else 0

            # Decrease the exploration rate at the end of every game
            self.qlearn_agent_black.decay_exploration_rate()
            self.qlearn_agent_white.decay_exploration_rate()

            # Add the results to the Excel sheet
            self.ws.append([self.current_game + 1, black_score, white_score, black_wins, white_wins,tie])

            print(f'Game {self.current_game + 1} has ended: BLACK {black_score}, WHITE {white_score}')
            self.display.display_results(self.results)

        if self.current_game < self.num_games:
            self.current_game += 1
            self.reset_game()
        else:
            self.finished = True
            print(self.finished)
            # Save the Excel file after all games are completed
            self.wb.save(str(self.results_file))
            print(f"end game and results saved to {self.results_file}")

    def run(self):
        self.play_game_step()


if __name__ == "__main__":
    size = 5
    num_games = 1000
    is_display = "display"

    root = tk.Tk()
    root.title("Go Game")

    display = GoDisplay(root, size)

    game = GoGameTrainingQLearning(size, num_games, display, is_display)
    print("start game")
    game.run()

    root.mainloop()
